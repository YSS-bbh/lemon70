# -*- coding: utf-8 -*-
"""
@Time �� 2020/7/27 20:14
@Author �� YaMeng
@File ��lesson7.py
@IDE ��PyCharm
@CopyRight������ʡ������Ϣ�������޹�˾
"""

'''
��������Ϊ��python�Զ�������Ӧ���Ǹ�ʲô���ӣ�
1��׼�����Զ����Ĳ�������    === done   test_case_api.xlsx
2��ʹ��pythonȥ��ȡ�������� === done   read_data()
4���������󣬵õ���Ӧ���    === done   api_func()
5��������жϣ� ִ�н�� vs  Ԥ�ڽ��  == ����
6���õ�һ�����ս������д����������   === done   write_result()
'''

'''
{'case_id': 3, 
'url': 'http://api.lemonban.com/futureloan/member/register', 
'data': '{"pwd":"12345678","type":1}', 
'expected': '{"code":1,"msg":"�ֻ���Ϊ��"}'}
'''
'''
���� a = '10 * 20'
eval(a) 
10 * 20
'''

import openpyxl
import requests


# ��ȡ��������
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row  # ȡ��sheet����������
    # print(max_row)
    case_list = []
    for i in range(2, max_row + 1, 1):  # ȡֵ��ȡ��ȡ�ң�����ҿ�
        dict1 = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,  # ȡ��url
            data=sheet.cell(row=i, column=6).value,  # ȡ��������
            expected=sheet.cell(row=i, column=7).value  # ȡ��Ԥ�ڽ��
        )
        case_list.append(dict1)  # dict1������һ��һ���Ĳ��������� --->װ���б�����  ����б�ʹ�������еĲ�������
    # print(case_list)
    return case_list


# ��������
def api_func(url, data):
    header_login = {'X-Lemonban-Media-Type': 'lemonban.v2',
                    'Content-Type': 'application/json'}
    res1 = requests.post(url=url, json=data, headers=header_login)
    # print(res1.json())
    response = res1.json()
    return response


# д����Խ��
def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)


# ��װ��һ��ִ�к���������
def execute_func(filename, sheetname):
    cases = read_data(filename, sheetname)  # ��ȡexcel��Ĳ�������
    # print(cases)
    for case in cases:  # ѭ��ȡ����������
        case_id = case.get('case_id')  # �ֵ�ȡֵ
        url = case['url']  # ȡurlͨ��key�ķ���
        data = case.get('data')  # ͨ��excelȡ����ֵ��str
        data = eval(data)  # eval() ���ã����б��ַ���������python���ʽ
        expected = case.get('expected')
        expected = eval(expected)
        expected_msg = expected.get('msg')  # ȡ��Ԥ�ڽ�����msg��Ϣ
        real_result = api_func(url=url, data=data)  # == �����˷�������ĺ��� ���� �������
        real_msg = real_result.get('msg')  # ȡ��ʵ��ִ�н�����msg��Ϣ
        print('Ԥ�ڽ��Ϊ��{}'.format(expected_msg))
        print('ʵ�ʽ��Ϊ��{}'.format(real_msg))
        if real_msg == expected_msg:
            print('��{}������ͨ����'.format(case_id))
            final_res = 'pass'
        else:
            print('��{}������δͨ����'.format(case_id))
            final_res = 'fail'
        print('*' * 30)
        write_result(filename, sheetname, case_id + 1, 8, final_res)


execute_func('test_case_api.xlsx', 'register')
execute_func('test_case_api.xlsx', 'login')
